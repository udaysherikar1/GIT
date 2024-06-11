import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import units
import os,csv

def read_csv(file_path):
    data = []
    with open(file_path, 'r') as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            data.append(row)
    return data
    
def createExcelReport():
    
    user_home = os.path.expanduser("~")
    #print('user_home --', user_home)
    image_csvPath = os.path.join(user_home, "temp_image_dir\captureImage.csv")
    #print('image_csvPath --', image_csvPath)
    csv_data=read_csv(image_csvPath)
    
    # Create a new workbook and select the active sheet
    #workbook = Workbook()
    template_Path = os.path.join(user_home, "temp_image_dir\ESTIMATION_TRACKING_SHEET_Template.xlsx")
    workbook = openpyxl.load_workbook(template_Path)
    sheet = workbook.active
    
    image_width = 400        # Adjust the width as needed
    image_height = 400       # Adjust the height as needed

    for row in csv_data:
       
        idx = int(row[0])
        name = row[1]
        image_path = row[2]

        i=idx+8
        cell_string = f'B{i}'
        #print('cell_string-->',cell_string)
        cell_image = f'C{i}'
        #print(cell_image)

        sheet[cell_string] = name
        img = Image(image_path)
        img.width = image_width
        img.height = image_height

        # Adjust the image position within the cell
        # Use the `units` module to convert pixels to EMUs (English Metric Units)
        #img.anchor = cell_image
        #img.anchor = f'C{i}:{units.pixels_to_EMU(500)}:{units.pixels_to_EMU(500)}'  # Offset 10 pixels from top-left

        #col, row = cell_image[:1], cell_image[1:]
        #img.anchor = f'{col}{row}:{10},{10}'
        
        # Calculate optimal row height based on image height
        row_height = image_height / 1.2  # Approximately 0.75 pixels per point (default Excel row height unit)
        sheet.row_dimensions[idx + 8].height = row_height
        
        sheet.add_image(img, cell_image)

        
        # Loop to add serial numbers in column A, names in column B, and images in column C
        # Add serial number to column A (starting from row 1)
        #cell_a = sheet.cell(row=idx + 1, column=1)
        #cell_a.value = idx
        #cell_a.alignment = Alignment(vertical='top')
        #sheet.cell(row=idx+8, column=1, value=idx)   # Write row number
        
        # Add name to column B (starting from row 1)
        #cell_b = sheet.cell(row=idx + 1, column=2)
        #cell_b.value = name
        #cell_b.alignment = Alignment(vertical='top')
        #sheet.cell(row=idx+8, column=2, value=name)  # Write name
        

        # Load image from the corresponding image path
        #img = Image(image_path)

        # Set the desired dimensions for the image (optional)
        #img_width = 200
        #img_height = 150
        #img.width = img_width
        #img.height = img_height

        # Calculate optimal row height based on image height
        #row_height = img_height / 1.2  # Approximately 0.75 pixels per point (default Excel row height unit)
        #sheet.row_dimensions[idx + 1].height = row_height

        # Add the image to the worksheet in column C (starting from row 1)
        # Calculate the top-left corner position of the image within the cell for proper alignment
        #cell_c = sheet.cell(row=idx + 1, column=3)
       #cell_c.alignment = Alignment(vertical='top')
        #sheet.add_image(img, cell_c.coordinate)



   
    # Save the workbook
    #image_csvPath = os.path.join(user_home, "temp_image_dir\excel_with_images.xlsx")
    #workbook.save('E:\\WorkingDir\\NEW_REPORT_AUTOMATION\\Models\\PT22\\dump\\excel_with_images.xlsx')
    #print('template_Path -->',template_Path)
    workbook.save(template_Path)
    
    
createExcelReport()    
